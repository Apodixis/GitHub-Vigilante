import os, openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from datetime import datetime

def _sanitize_excel_value(value):
	"""Remove control characters that openpyxl cannot store in worksheet cells."""
	if isinstance(value, str):
		return ILLEGAL_CHARACTERS_RE.sub("", value)
	return value

def write_to_excel(user_data, target, search_mode) -> str:
	"""
	Writes search results (without enrichment) to an Excel file.
	The file is named as: f"{YYYYMMDDHHMM}{search_mode}_{target}.xlsx"
	data: outer dict key values used as column headers, outer dict field values used in corresponding column cell values
	"""
	# Create workbook and worksheet
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = search_mode
    
	# Preserve column order: start with keys from first user, append any new keys found in others
	if user_data:
		all_keys = list(user_data[0].keys())
		for user in user_data[1:]:
			for k in user.keys():
				if k not in all_keys:
					all_keys.append(k)
	else:
		all_keys = []
    
	# Write header
	for col, key in enumerate(all_keys, 1):
		ws.cell(row=1, column=col, value=key)
    
	# Write user data
	for row, user in enumerate(user_data, 2):
		for col, key in enumerate(all_keys, 1):
			val = user.get(key, "")
			
			# Convert sets/lists to comma-separated string for Excel
			if isinstance(val, (set, list)):
				val = ', '.join(str(item) for item in val)
			
			elif isinstance(val, dict):
				val = str(val)
			
			val = _sanitize_excel_value(val)
			ws.cell(row=row, column=col, value=val)
    
	# Autosize columns
	for col in range(1, len(all_keys)+1):
		ws.column_dimensions[get_column_letter(col)].auto_size = True
    
	# Build filename and path to Downloads
	date_str = datetime.now().strftime("%Y%m%d%H%M")
	filename = f"{date_str}{search_mode}_{target}.xlsx"
	downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
	file_path = os.path.join(downloads_folder, filename)
    
    # Save results to workbook
	wb.save(file_path)
	return filename